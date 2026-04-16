#!/usr/bin/env python3
"""Download e2e4 ZIP, extract and flatten workbooks into one CSV.

This script combines download, extraction and flattening into a single command.

Usage:
  python scripts/e2e4_ingest.py
  python scripts/e2e4_ingest.py --sample 100
  python scripts/e2e4_ingest.py --force
"""
from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
import zipfile
from pathlib import Path

import requests
from openpyxl import load_workbook


DEFAULT_URL = "https://e2e4online.ru/ws/excel/irkutsk.e2e4online.ru.zip"


def safe_extract(zip_path: Path, target_dir: Path, overwrite: bool = False) -> None:
    with zipfile.ZipFile(zip_path, "r") as z:
        target_dir.mkdir(parents=True, exist_ok=True)
        abs_target = str(target_dir.resolve())
        for member in z.infolist():
            member_name = member.filename
            if member_name.endswith("/") or member_name.endswith("\\"):
                continue
            dest_path = target_dir.joinpath(member_name)
            dest_path_parent = dest_path.parent
            resolved = dest_path.resolve()
            if not str(resolved).startswith(abs_target):
                raise RuntimeError(f"Unsafe zip member path: {member_name}")
            dest_path_parent.mkdir(parents=True, exist_ok=True)
            if dest_path.exists() and not overwrite:
                # skip existing files
                continue
            with z.open(member) as src, open(dest_path, "wb") as dst:
                shutil.copyfileobj(src, dst)


def download(url: str, dest: Path, force: bool = False) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and not force:
        print(f"ZIP already exists: {dest}")
        return dest
    print(f"Downloading: {url}")
    with requests.get(url, stream=True, timeout=60) as r:
        r.raise_for_status()
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
    return dest


def looks_cyr(s: str) -> int:
    return len(re.findall(r"[а-яёА-ЯЁ]", s))


def fix_mojibake(s: object) -> str:
    if s is None:
        return ''
    s = str(s)
    if looks_cyr(s) > 0 and '�' not in s:
        return s.strip()
    try:
        b = s.encode('latin-1', errors='ignore')
        for enc in ('cp1251', 'cp866'):
            try:
                cand = b.decode(enc)
                if looks_cyr(cand) > 0:
                    return cand.strip()
            except Exception:
                continue
    except Exception:
        pass
    return s.strip()


def detect_header(sheet, max_rows=40):
    best_row_idx = None
    best_count = -1
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        vals = [v for v in row if v is not None and str(v).strip() != '']
        if len(vals) > best_count:
            best_count = len(vals)
            best_row_idx = i
    return best_row_idx


def normalize_header(raw_header):
    hdr = [fix_mojibake(x) for x in raw_header]
    out = []
    seen = {}
    for i, h in enumerate(hdr):
        name = h or f"col_{i}"
        name = re.sub(r"\s+", " ", name).strip()
        if not name:
            name = f"col_{i}"
        key = name
        if key in seen:
            seen[key] += 1
            key = f"{name} {seen[key]}"
        else:
            seen[key] = 1
        out.append(key)
    return out


def flatten_workbook(xlsx_path: Path, out_csv_path: Path, sample_limit: int | None = None) -> int:
    wb = load_workbook(filename=str(xlsx_path), read_only=True)
    try:
        out_csv_path.parent.mkdir(parents=True, exist_ok=True)
        total_written = 0
        with out_csv_path.open('w', newline='', encoding='utf-8-sig') as csvf:
            writer = None
            for sheetname in wb.sheetnames:
                sh = wb[sheetname]
                header_idx = detect_header(sh, max_rows=40)
                if header_idx is None:
                    continue
                header_row = None
                for i, row in enumerate(sh.iter_rows(values_only=True)):
                    if i == header_idx:
                        header_row = row
                        break
                if header_row is None:
                    continue
                columns = normalize_header(header_row)
                if writer is None:
                    fieldnames = ['source_file', 'sheet'] + columns
                    writer = csv.DictWriter(csvf, fieldnames=fieldnames, delimiter=';')
                    writer.writeheader()
                empty_streak = 0
                for i, row in enumerate(sh.iter_rows(values_only=True)):
                    if i <= header_idx:
                        continue
                    row_vals = [fix_mojibake(v) for v in row]
                    if all((not (v and str(v).strip())) for v in row_vals):
                        empty_streak += 1
                        if empty_streak >= 20:
                            break
                        else:
                            continue
                    empty_streak = 0
                    d = {'source_file': xlsx_path.name, 'sheet': sheetname}
                    for j, col in enumerate(columns):
                        d[col] = row_vals[j] if j < len(row_vals) else ''
                    writer.writerow(d)
                    total_written += 1
                    if sample_limit and total_written >= sample_limit:
                        return total_written
        return total_written
    finally:
        try:
            wb.close()
        except Exception:
            pass


def main(argv=None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('--url', default=DEFAULT_URL)
    parser.add_argument('--out-dir', default=str(Path(__file__).resolve().parent.parent / 'data' / 'e2e4'))
    parser.add_argument('--downloads', default=str(Path(__file__).resolve().parent.parent / 'data' / 'downloads'))
    parser.add_argument('--out-csv', default=str(Path(__file__).resolve().parent.parent / 'data' / 'catalogs' / 'e2e4_flat.csv'))
    parser.add_argument('--force', action='store_true')
    parser.add_argument('--no-download', action='store_true')
    parser.add_argument('--sample', type=int, default=0, help='Stop after N rows (quick test)')
    args = parser.parse_args(argv)

    out_dir = Path(args.out_dir)
    downloads = Path(args.downloads)
    downloads.mkdir(parents=True, exist_ok=True)
    zip_name = Path(args.url).name or 'e2e4.zip'
    zip_path = downloads / zip_name

    if not args.no_download:
        try:
            download(args.url, zip_path, force=args.force)
        except Exception as e:
            print('Download failed:', e)
            return 2
    else:
        if not zip_path.exists():
            print('ZIP not found:', zip_path)
            return 3

    try:
        safe_extract(zip_path, out_dir, overwrite=args.force)
    except Exception as e:
        print('Extract failed:', e)
        return 4

    # remove zip archive after successful extraction
    try:
        if zip_path.exists():
            zip_path.unlink()
            print(f"Deleted zip archive: {zip_path.name}")
    except Exception as e:
        print(f"Failed to delete zip archive: {e}")

    files = sorted(out_dir.glob('*.xls*'))
    if not files:
        print('No workbook files found in', out_dir)
        return 5
    out_csv = Path(args.out_csv)
    total = 0
    for f in files:
        print('Processing', f.name)
        written = flatten_workbook(f, out_csv, sample_limit=(args.sample or None))
        print('  wrote rows:', written)
        total += written
        # delete source xlsx after flattening
        try:
            if f.exists():
                f.unlink()
                print(f"  deleted source workbook: {f.name}")
        except Exception as e:
            print(f"  failed to delete {f.name}: {e}")
        if args.sample and total >= args.sample:
            break

    print('Done. total rows written:', total, '->', out_csv)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
